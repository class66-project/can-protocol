"""
/* Copyright (C) 2024 Richard Franks - All Rights Reserved
 *
 * You may use, distribute and modify this code under the
 * terms of the Apache 2.0 license.
 *
 * See LICENSE for details
 */
"""
import os
from openpyxl import load_workbook

HEADER = """
/* Copyright (C) 2024 Richard Franks - All Rights Reserved
 *
 * You may use, distribute and modify this code under the
 * terms of the Apache 2.0 license.
 *
 * See LICENSE for details
 */

#pragma once
"""

def run(filename):
    """Extracts the config addresses from the specified file"""
    workbook = load_workbook(filename=filename, read_only=True)

    defines = []

    max_length = 0

    for row in workbook["Config Addresses"].iter_rows(
        min_row = 1,
        max_row = 1500,
        min_col = 1,
        max_col = 6,
        values_only = True):

        parts = list(row)
        if None in (parts[1], parts[3], parts[4]):
            continue
        if parts[1][:2] != "0x":
            continue

        if parts[2] == "" or parts[3] == "":
            continue

        define = {}
        define["value"] = ""
        define["name"] = ""
        define["size"] = ""
        define["units"] = ""
        define["notes"] = ""

        name = parts[3].upper().replace(" ", "_")

        max_length = max(len(name), max_length)

        define["value"] = parts[1]
        define["name"] = name
        define["size"] = parts[4]
        defines.append(define)

    max_length += 2
    lines = []
    for define in defines:
        if define["value"][-1:] == "0":
            lines.append("")
        line = f"#define {define['name'].ljust((max_length + max_length % 2))}{define['value']}"
        line += f"  // Size: {define['size']}"
        lines.append(line)

    file_path = os.path.dirname(os.path.realpath(__file__))
    with open(f"{file_path}/../config.h", "w", encoding="utf-8") as fh:
        fh.write(HEADER)
        fh.write("\n".join(lines))
        fh.write("\n")

if __name__ == "__main__":
    run("/mnt/c/Users/Richard/OneDrive/Documents/66Shed/Can Protocol.xlsx")
