import os
import sys
import logging
import subprocess
from openpyxl import load_workbook

logging.basicConfig(level=logging.DEBUG)

header = """
/* Copyright (C) 2024 Richard Franks - All Rights Reserved
 *
 * You may use, distribute and modify this code under the
 * terms of the Apache 2.0 license.
 *
 * See LICENSE for details
 */

#pragma once
"""

def run(filename):
    workbook = load_workbook(filename=filename, read_only=True)

    defines = []

    maxLength = 0

    for row in workbook["Sensor Addresses"].iter_rows(min_row = 1, max_row = 300, min_col = 1, max_col = 7, values_only = True):
        parts = list(row)
        if None in (parts[1], parts[3], parts[4]):
            continue
        if parts[1][:2] != "0x":
            continue
        if parts[1] == "0x00":
            # skipping reserved address
            continue
        
        if parts[2] == "" or parts[3] == "":
            continue
        
        define = {}
        define["value"] = ""
        define["name"] = ""
        define["size"] = ""
        define["units"] = ""
        define["notes"] = ""

        name = ""
        if parts[2][:1] != "!":
            name += f"{parts[2].upper()}_"
        
        name += parts[3].upper().replace(" ", "_")

        maxLength = max(len(name), maxLength)

        define["value"] = parts[1]
        define["name"] = name
        define["size"] = parts[4]
        define["units"] = parts[5] or ''
        defines.append(define)
    
    maxLength += 2
    lines = []
    for define in defines:
        if define["value"][-1:] == "0":
            lines.append("")
        lines.append(f"#define {define['name'].ljust((maxLength + maxLength % 2))}{define['value']}  // Size: {define['size']}")

    with open("../sensors.h", "w") as fh:
        fh.write(header)
        fh.write("\n".join(lines))

if __name__ == "__main__":
    run("/mnt/c/Users/Richard/OneDrive/Documents/66Shed/Can Protocol.xlsx")